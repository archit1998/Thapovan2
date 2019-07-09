


import pandas as pd
import xlrd
import numpy as np
import openpyxl as opx
import glob
import sys
import datetime
import xlwings as xl
import copy
from openpyxl.utils import range_boundaries
from fuzzywuzzy import fuzz
import pyodbc
import math as math
path = 'E:\\trial'
ticker_list =[] 
folder = glob.glob(path + '\\*')
company = [1]
import ast
c = 0
import configparser
con=configparser.ConfigParser()
con.read("E:\\final\\Config.cfg")


# In[22]:


import pandas as pd
import os
import gensim
from gensim.utils import simple_preprocess
from gensim.parsing.preprocessing import STOPWORDS
from nltk.stem import WordNetLemmatizer, SnowballStemmer
from nltk.stem.porter import *
from nltk import PorterStemmer
stemmer = PorterStemmer()
import numpy as np
np.random.seed(2018)
import nltk
from nltk.stem.porter import PorterStemmer
from nltk.stem.wordnet import WordNetLemmatizer
nltk.download('wordnet')
lem = WordNetLemmatizer()
stem = PorterStemmer()
import re
import nltk
nltk.download('stopwords')
from nltk.corpus import stopwords
from nltk.stem.porter import PorterStemmer
from nltk.tokenize import RegexpTokenizer
from nltk.stem.wordnet import WordNetLemmatizer
stop_words = set(stopwords.words("english"))
new_words = ["using", "show", "result", "large", "also", "iv", "one", "two", "new", "previously", "shown"]
stop_words = stop_words.union(new_words)
import time
from tabulate import tabulate


# In[23]:


os.chdir("E:\\final\\Corpus")
corpus_list = glob.glob("E:\\final\\Corpus\\*csv")
excel_files =  glob.glob("E:\\trial\\2018\\may\\2 MAY\\lvs\\*xlsx")


# In[24]:


def safe_run(func):

    def func_wrapper(*args, **kwargs):

        try:
           return func(*args, **kwargs)

        except OperationalError as e:
            server = ast.literal_eval(con.get("Setting","server"))
            database = ast.literal_eval(con.get("Setting","database"))
            username = ast.literal_eval(con.get("Setting","username"))
            password = ast.literal_eval(con.get("Setting","password"))
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            cursor = cnxn.cursor()
            return func(*args, **kwargs)

    return func_wrapper


# In[25]:


def spinning_cursor():
  #while True:
    for cursor in '\\-/|':
      time.sleep(0.3)
      # Use '\r' to move cursor back to line beginning
      # Or use '\b' to erase the last character
      sys.stdout.write('\r{}'.format(cursor))
      # Force Python to write data into terminal.


# In[26]:


data1 = pd.read_csv("E:\\Thapovan2\\Corpus\\values1.csv", error_bad_lines=False)
data = pd.read_csv("E:\\Thapovan2\\Corpus\\Thapovan_data.csv", error_bad_lines=False)


# In[ ]:





# In[ ]:





# In[ ]:





# In[27]:


def p():
    pass


# In[28]:


#connection to database:
server = ast.literal_eval(con.get("Setting","server"))
database = ast.literal_eval(con.get("Setting","database"))
username = ast.literal_eval(con.get("Setting","username"))
password = ast.literal_eval(con.get("Setting","password"))
cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
cursor = cnxn.cursor()


# In[29]:



data = []
#function which get information of file using its path and return 5 elements(year, month,date,ticker code,analyst name)
def formating(filen):
    data = filen.replace(path, '').split('\\')
    del data[0]
    del data[3]
    if '-' in data[-1]:
        temp = (data[-1].split('.')[0]).split('-')
    elif '_' in data[-1]:
        temp = (data[-1].split('.')[0]).split('_')
    else:
        temp = (data[-1].split('.')[0]).split(' ')
    del data[-1]
    data.append(temp[0])
    data.append(temp[1])
    print(data)
    return(data)


# In[30]:



#defined variable which help us to recognize year and quater of given column
year_pattern_int=[]
year_pattern_float=[]
year_pattern_2di=[]
quater_pattern=ast.literal_eval(con.get("Setting","quater_pattern"))
bscf_names=ast.literal_eval(con.get("Setting","bscf_names"))
#bscf_names=["balancesheet","cashflows","bs&cf","cf&bs","bsheet","balancesheet&cashflow","cashflow&balance sheet","cash&balance","adjustedbs","annualbs&cf","annualcf&bs"]

check_list = ast.literal_eval(con.get("Setting","check_list"))
for i in range(2000,2031):
    year_pattern_int.append(i)
    year_pattern_float.append(float(i))
    if i-2000<10:
        year_pattern_2di.append('0'+str(i-2000))
    else:
        year_pattern_2di.append(str(i-2000))


# In[31]:


# type of main function
#@safe_run
def function1(excel_file,path_list):
    #read excel file using openpyxl
    wb=opx.load_workbook(excel_file,data_only=True)
    #print("Hi")
    pd.set_option('expand_frame_repr', False)
    sheet_tobechecked=wb.sheetnames #getting all the sheet name
    nonhidden=[]
    startrow={}
    print("\nAll the sheets prestent in the file are: ")
    print(sheet_tobechecked)
    #this part helps to ignore hidden sheets and to take related to BS , Cf and IS
    names=[]
    for i in range(0, len(sheet_tobechecked)):
        curr_sheet = wb[sheet_tobechecked[i]]
        if curr_sheet.sheet_state != 'hidden':
            nonhidden.append(sheet_tobechecked[i])
            if sheet_tobechecked[i].lower()=='bs' or sheet_tobechecked[i].lower()=='cf' or sheet_tobechecked[i].lower()=='model':
                names.append(sheet_tobechecked[i])
            else:
                count_nobscf=1
                for nam in bscf_names:
                    if fuzz.partial_token_set_ratio(sheet_tobechecked[i].lower(),nam)>=80:
                        names.append(sheet_tobechecked[i])
    
    sheet_tobechecked = names # all the sheet which are important to us
    
    #if no sheet found then return giving error as sheet_name_fault
    if len(sheet_tobechecked)<=0:
        return (2) # that means there is problem
        '''print(nonhidden)
        while True:
            sheee=input("Enter the sheet name U want to process from above or for exit input'e' ")
            if sheee!='e':
                sheet_tobechecked.append(sheee)
            else:
                break'''
    
    sheet_tobechecked=list(set(sheet_tobechecked))# to remove repeated sheet names
    print("Sheets to be checked: ")
    #print(sheet_tobechecked)
    #for i in range(0,len(sheet_tobechecked)):
       # print(str(i) + ") " + sheet_tobechecked[i])
    tst = 0
    while tst == 0:
        for i in range(0,len(sheet_tobechecked)):
            print(str(i) + ") " + sheet_tobechecked[i])
        a=input("if you want to continue with these sheets, type [Y]es. If you want to remove some sheets, type the corresponding number seperated by a comma(,)")
        if(a=="Y"):
            tst == 1
            break
        else:
            t =[]
            t=a.split(",")
            t1 =[]
        for i in range(0,len(t)):
            t[i] = int(t[i])
            t1.append(sheet_tobechecked[t[i]])
        for i in t1:
            if i in sheet_tobechecked:
                sheet_tobechecked.remove(i)
    #defining variable
    year_pre,quater_pre,q4_check_pre='','',0
    
    #unmerging all cells of every sheet of a file
    for every_sheet in sheet_tobechecked:
        
        startrow_no=startingrow(wb,every_sheet)-1
        if startrow_no<0:
            startrow[every_sheet]=0
        else:
            startrow[every_sheet]=startrow_no
        
        curr_sheet=wb[every_sheet]
        merge_list=copy.deepcopy(curr_sheet.merged_cells.ranges)
        for group in merge_list :
            split_arr = group.coord.split(':')
            min_col, min_row, max_col, max_row = range_boundaries(group.coord)
            top_left_cell_value = curr_sheet[split_arr[0]].value
            curr_sheet.unmerge_cells(group.coord)
            for row in curr_sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value
    wb.save(excel_file) #the modified file is been overwritten (have to look over it afterward by changing path)
    #print("excel")
    #display(excel_file)
    for sheetno in range(0,len(sheet_tobechecked)):
        print(sheet_tobechecked[sheetno])
        #print(sheet_tobechecked[sheetno])
        address_dataframe=pd.DataFrame(columns=['q1','q2','q3','q4','tl','Ex']) # defining lookup table
        excel_data = pd.read_excel(excel_file,sheet_name=sheet_tobechecked[sheetno],header=None,skiprows=[i for i in range(0,startrow[sheet_tobechecked[sheetno]])])
        pd.options.display.max_columns = None
        pd.options.display.max_rows=None
        #print("This is excel data")
        #display(excel_data)
        
        #formating the dataframe by removing column and row having less than 10 element
        excel_data=excel_data.replace(0,np.nan)
        excel_data=excel_data.dropna(axis=0,thresh=10)
        excel_data=excel_data.dropna(axis=1,thresh=10)
        excel_data=excel_data.reset_index(drop=True)
        if(len(excel_data.columns)==0):
            continue # this line was added to handle exceptions in the code
        excel_data.columns = range(excel_data.shape[1])
        pd.options.display.max_columns = None
        pd.options.display.max_rows=None
        #display(excel_data)
        startrowno=0
        print("..........10%")
        #getting defination of each column
        #print(len(excel_data.columns))
        if len(excel_data.columns)>10: #changed from 10 to 3
            temp_dataframe=excel_data[:10] # the first 10 element of a column will tell us about it 
            for j in temp_dataframe:
                list_ele=list(temp_dataframe[j])
                knowledge=[]
                colno=j
                for i in list_ele:
                    #overlook element if its nan,none, or weeks or days written in it
                    if str(i)=='nan' or str(i)=='None' or type(i)==datetime.datetime or "week" in str(i) or "day" in str(i):
                        a=1
                    #take only element which are keywords to us (like FY2017,3Q17 etc.) or else break which means any value have been occured
                    elif (type(i)==float and i not in year_pattern_float) or (type(i)==int and i not in year_pattern_int):
                        startrowno=list_ele.index(i)
                        break
                    else:
                        knowledge.append(i)
                
                # proces further if the list taken for defining the row is not empty 
                if len(knowledge)>0:
                    
                    year_got,quater_got,q4_check_got=about_q_year(knowledge)
                        
                    #here is a small assumtion (if we got Q4 and all the elementused to define column is same we will take the curent row as Total of particuar year) )
                    if q4_check_got==1 and year_got==year_pre and quater_got==quater_pre and q4_check_got==q4_check_pre:
                        year_final=year_got
                        quater_final='TL'
                        q4_check_final=1
                    else:
                        # or else everything will be consider as it was given has been found about it
                        year_pre=year_got
                        quater_pre=quater_got
                        q4_check_pre=q4_check_got
                        year_final=year_got
                        quater_final=quater_got
                        q4_check_final=q4_check_got                       
                    
                    #if only year is written and nothing about the quater we take that as Total for particular year
                    if year_got!='' and (quater_got=='' or quater_got=='TL' or quater_pre=='TL' or quater_pre==''):
                        quater_final='TL'
                    
                    #printing and making dataframe(lookup table) of information of each columns
                    if year_final!='' and quater_final!='':
                        got_cor=1 #means we got something
                        #print(knowledge,"===",end="")
                        #print("belongs to "+"20"+year_final," "+quater_final)
                        if '1' in quater_final:
                            qval='q1'
                        elif '2' in quater_final:
                            qval='q2'
                        elif '3' in quater_final:
                            qval='q3'
                        elif '4' in quater_final:
                            qval='q4'
                        else:
                            qval='tl'
                        try:
                            address_dataframe.loc["20"+year_final][qval]=[startrowno,colno]
                        except Exception:
                            address_dataframe.loc["20"+year_final]='Nan'
                            address_dataframe.loc["20"+year_final][qval]=[startrowno,colno]                           
                 
        #display(address_dataframe)
        print("..........20%")
        #db_connect()
        #extra fro shoeing rows which doesnot match
        not_present=pd.DataFrame(columns=excel_data.columns)
        coll=0
        
        #this is to make a list of year and its respective column number from lookup take which has to consider
        row_name_year=list(address_dataframe.index)
        row_name_year.sort()
        print(row_name_year)
        if int(path_list[0])<int(row_name_year[-1]):
            cur_yr=int(path_list[0])
            count_yr=0
            count = 0
            metric_col=[]
            par_year=[]
            metric_row=20
            year_tosearch=cur_yr
            
            lowest_year=min_year("'"+path_list[3]+"'")

            while (count_yr<6 and year_tosearch>=lowest_year):
                year_tosearch=cur_yr-(count_yr+1)
                val=address_dataframe['q1'][str(year_tosearch)]
                if val!='Nan':
                    if val[0]<metric_row:
                        metric_row=val[0]
                    metric_col.append(val[1])
                    par_year.append(year_tosearch)
                    count_yr+=1
            print("Years considered:")
            print(par_year)
            print("Columns used:")
            print(metric_col)
            print("Start Row: ")
            print(metric_row)
            print("..........30%")
                

            all_metric=set()
            def_metrics = []
            matched_metric = []
            mid_matched =[]
            maybe_metrics = []
            no_metrics = []
            if(type(excel_data.iloc[metric_row][0])==str):
                c = 0
            else:
                c = 1    
            metric_list =[]    
            list_subs_found =[]
            for rowss in range(metric_row,excel_data.shape[0]):
                metric_found=[]
                s_var = ['Las Vegas - Venetian & Palazzo', 'Sands Bethworks, PA', 'Marina Bay Sands Singapore', 'Venetian (including Palazzo)', 'Venetian Macao', 'Sands Macao', 'Plaza Casino & Four Seasons Macao', 'Sands Cotai Central', 'Parisian', 'Marina Bay Sands', 'Sands Bethworks', 'Other Asia and Eliminations', 'Other Asia', 'Venetian/Palazzo, Las Vegas', 'Venetian, Macau', 'Sands Macau,', 'Four Seasons, Macau', 'Cotai Central, Macau', 'The Parisian, Macau', 'Marina Bay Sands, Singapore', 'Bethworks, Bethlehem, PA', 'Other', 'EBITDAR (Before Corporate Expense)', 'Sands Macau', '1. Venetian + Palazzo Las Vegas', '3. Venetian Macao', '2. Sands Macao', '4. Four Seasons Macao', '5. Cotai Central', '6. Parisian Macao', '7. Marina Bay Sands (Singapore)', '9. Other, Net', 'Las Vegas Operations', 'Venetian Macau', 'Four Seasons Macau', 'Cotai Central', 'The Parisian Macao', 'Maina Bay - Singapore', 'TOTAL PROPERTY EBITDAR', 'Marina Bay - Singapore', 'Parisian,', 'Las Vegas Operating Properties', 'The Venetian Macau', 'Singapore', 'Venetian / Palazzo', 'Sands Macao ', 'Four Seasons Macao ', 'Sands Cotai Central (Sites 5&6) ', 'Parisian (Site 3) ', 'Marina Bay Sands ', 'Other Asia ', 'Las Vegas', 'Macao', 'Total Adj. Property EBITDA', 'Las Vegas Operations (Venetion LV+ Palazzo+Sands Expo)', 'The Venetian Macao', 'The Plaza Macao/Four Seasons', 'The Marina Bay Sands', 'Las Vegas Properties', 'Total Parisian Net Revs', 'Macau Property Revenues (Ex. Ferries)', 'Four Seasons Hotel Macao', 'Las Vegas Property', 'Four Seasons- Cotai', 'The Parisian', 'Las Vegas Properities', 'Cotai Plot 2 (Four Seasons)', 'Cotai Plot 5 & 6 (Shangri-La/Traders/Sheraton)', 'Marina Bay Sands (Singapore)']
                list_metric_found=[]
                com_met=set()
                a_metric=str(excel_data.iloc[rowss][c])
                flag_sub = 0
                for i in s_var:
                    if i in a_metric:
                        s = i
                        flag_sub = 1
                if flag_sub==0:
                    if rowss>metric_row+5:
                        for i in range(1,5):
                            for j in s_var:
                                if j in str(excel_data.iloc[rowss-i][c]):
                                    s=j
                                    flag_sub =1
                            if flag_sub==1:    
                                    break
                if flag_sub==0: 
                    s = "Unknown"
                for met_num in range(0,6):
                    #getting the metric name wrt the year,quater,ticker and value send
                    metric,subs_list=get_metric_name(par_year[met_num],path_list[3],excel_data.iloc[rowss][metric_col[met_num]],a_metric)
                    #remove this if it fucks up
                    for i in range(0,len(metric)):
                        metric[i] = metric[i] +"["+str(subs_list[i])+"]"
                    metric = set(metric)
                    subs_list = set(subs_list)
                    com_met.update(metric)
                    metric_list.extend(metric)
                    list_metric_found.extend(metric)
                    list_subs_found.extend(subs_list)
                if com_met:
                    spinning_cursor()
                    #print(com_met)
                    #print(list_metric_found)
                for un_met in com_met:
                    #if min 60% of them says that yes the values are of  particlar metric we take it 
                    if list_metric_found.count(un_met)/6 >= 0.6:
                        a = un_met +"--->"+a_metric+ "--->"+ str(rowss) +"--->"+ s
                        metric_found.append(un_met)
                        matched_metric.append(un_met)
                        def_metrics.append(a)
                    elif list_metric_found.count(un_met)/6 >= 0.4 and list_metric_found.count(un_met)/6 <0.6:
                        if(un_met not in matched_metric): 
                            mid_matched.append(un_met)
                            a = un_met +"--->"+a_metric+ "--->"+ str(rowss) +"--->"+ s
                            maybe_metrics.append(a)
                    elif list_metric_found.count(un_met)/6 < 0.4:
                        if(un_met not in matched_metric and un_met not in mid_matched):
                            no_metrics.append(un_met)
                if metric_found:
                    #print(rowss,metric_found)
                    all_metric.update(metric_found)
                else:
                    not_present.loc[coll]=list(excel_data.loc[rowss])
                    coll+=1
                
            #print(all_metric,len(all_metric))#printing all mertric found in database 
            print("..........90%")
            #metric_db=get_metriclist_from_db(path_list[0],path_list[3])
            #print("metric not in actuals---")
            #print(list(set(metric_db)-set(all_metric)))# all metric we couldn't find wrt to a ticker
            #print("metric not with respect to analyst")
            #print(list(set(get_metriclist_wrt_analyst(path_list[3],path_list[4]))-set(all_metric)))#metric which we couldn't finf wrt to analyst
            # the percentge accuracy will be calculate wby no of metric_found by us divided by total number of metric wrt analyst 
            #print()
            #print()
            #print("ACCURACY==")
            #all_metric = list(all_metric)
            #print("Number of total element in Actual DB=", len(metric_db))
            #print("Number of metric found by us=",len(all_metric))
            #for i in range(0,len(all_metric)):
            #    all_metric[i] = all_metric[i].replace(all_metric[i][all_metric[i].find("[")+1:all_metric[i].find("]")],"").strip("][")
            #print("accuracy wrt to Actual=",len(list(set(metric_db)&set(all_metric)))*100/len(metric_db))
            #print()
            #print("Number of total element in Estimated DB wrt Analyst=", len(get_metriclist_wrt_analyst(path_list[3],path_list[4])))
            #z=set(get_metriclist_wrt_analyst(path_list[3],path_list[4]))
            #print("Number of metric found by us=",len(z&set(all_metric)))
            #print("Accuracy with respect to estimates=",len(z&set(all_metric))*100/len(z))
            print("\n\nlist of definate metrics \n", list(set(def_metrics)))
            print()
            print("\n\nlist of semi-definate metrics \n", list(set(maybe_metrics)))
            print()
            print("\n\nlist of Unmatched metrics \n", list(set(no_metrics)))
            m_metrics = list(set(maybe_metrics))
            d_metrics = list(set(def_metrics))
            print()
            print("\n\nThe partially matched metrics can be mapped as so")
            print()
            for metric in m_metrics:
                poss = []
                poss = metric.split("--->")
                col=[c]
                for i in range(0,6):
                    col.append(metric_col[i])
                temp_metric = pd.DataFrame()
                temp_metric=temp_metric.append(excel_data.loc[[int(poss[2])],col[0:7]])
                temp_metric.rename(columns={col[0]:'Name',col[1]:str(par_year[0])+'Q1',col[2]:str(par_year[1])+'Q1',col[3]:str(par_year[2])+'Q1',col[4]:str(par_year[3])+'Q1',col[5]:str(par_year[4])+'Q1',col[6]:str(par_year[5])+'Q1'},inplace=True)
                yr = str(par_year[0:6]).strip('][').replace(" ,","")
                #display(temp_metric)
                sub = int(poss[0][poss[0].find("[")+1:poss[0].find("]")])
                poss0 = poss[0].replace(poss[0][poss[0].find("[")+1:poss[0].find("]")],"").strip("][")
                select_stmt1 = F"select value from dbo.cc_actual_metrics_consolidated where "                     F"metric_name='{poss0}' and subsidiary_id ={sub} and quarter = 'Q1' and ticker_code ='{path_list[3]}' and financial_year in (select convert(int,id) from [adm_fnSplitter]('{yr}')) order by financial_year"  
                cursor.execute(select_stmt1)
                row = cursor.fetchone()
                m_l =[]
                m_l1 = [poss0]
                while row:
                    m_l.append(str(row.value))
                    row= cursor.fetchone()
                m_l.reverse()
                m_l1.extend(m_l)
                if(len(m_l1)<7):
                    n = 7-len(m_l1)
                    for i in range(0,n):
                        m_l1.append("None")
                print(m_l1)
                temp_metric = temp_metric.append(pd.Series(m_l1, index=['Name',str(par_year[0])+'Q1',str(par_year[1])+'Q1',str(par_year[2])+'Q1',str(par_year[3])+'Q1',str(par_year[4])+'Q1',str(par_year[5])+'Q1']), ignore_index=True) 
                types = ["Analyst metric","Database metric"]
                temp_metric['Type'] = types
                #temp_metric=temp_metric.set_index('Type')
                #display(temp_metric)
                print(tabulate(temp_metric, headers='keys', tablefmt='psql'))
                #print(m_l)  
                #print(f"{'DB metric name and subsidiary id =':<30}{poss[0]:>15}",
                      #f"\n{'Analyst metric name =':<30}{poss[1]:>15}",
                      #f"\n{'Row Number in excel =':<30}{poss[2]:>15}",
                      #f"\n{'Row Number in excel =':<30}{poss[3]:>15}")
                print("DB metric name and subsidiary id: \n" + poss[0])
                print("Analyst metric name: \n" + poss[1])
                print("Row Number in excel: \n" + poss[2])
                print("Subsidiary Name from Analyst: \n" + poss[3])
                ip = input("\nenter [Y]es if you think this is a valid match and [N]o if you think it can be bypassed")
                if(ip is "Y"):
                    def_metrics.append(metric)
                else:
                    continue
            #matched_df = pd.DataFrame()
            print("\n\nthe Def matched metrics are")
            for m in d_metrics:
                poss = []
                poss = m.split("--->")
                col=[c]
                an_metric = pd.DataFrame(columns = ['Name',str(par_year[0])+'Q1',str(par_year[1])+'Q1',str(par_year[2])+'Q1',str(par_year[3])+'Q1',str(par_year[4])+'Q1',str(par_year[5])+'Q1'])
                temp_metric = pd.DataFrame()
                columns = ['Name',str(par_year[0])+'Q1',str(par_year[1])+'Q1',str(par_year[2])+'Q1',str(par_year[3])+'Q1',str(par_year[4])+'Q1',str(par_year[5])+'Q1']
                for i in range(0,6):
                    col.append(metric_col[i])
                temp_metric=temp_metric.append(excel_data.loc[[int(poss[2])],col[0:7]])
                temp_metric.rename(columns={col[0]:'Name',col[1]:str(par_year[0])+'Q1',col[2]:str(par_year[1])+'Q1',col[3]:str(par_year[2])+'Q1',col[4]:str(par_year[3])+'Q1',col[5]:str(par_year[4])+'Q1',col[6]:str(par_year[5])+'Q1'},inplace=True)
                #display(temp_metric)  
                inx = metric_list.index(poss[0])
                sub = int(poss[0][poss[0].find("[")+1:poss[0].find("]")])
                poss0 = poss[0].replace(poss[0][poss[0].find("[")+1:poss[0].find("]")],"").strip("][")
                #print(sub)
                if sub is not None and sub is not "None":    
                    select_stmt1 = F"select value from dbo.cc_actual_metrics_consolidated where "                         F"metric_name = '{poss0}' and subsidiary_id = {sub} and ticker_code = '{path_list[3]}' and quarter = 'Q1' and financial_year in (select convert(int,id) from [adm_fnSplitter]('2012,2013,2014,2015,2016,2017')) order by financial_year"
                else:
                    select_stmt1 = F"select value from dbo.cc_actual_metrics_consolidated where "                         F"metric_name = '{poss0}' and ticker_code = '{path_list[3]}' and quarter = 'Q1' and financial_year in (select convert(int,id) from [adm_fnSplitter]('2012,2013,2014,2015,2016,2017')) order by financial_year"
                cursor.execute(select_stmt1)
                row = cursor.fetchone()
                m_l =[]
                m_l1 = [poss0]
                while row:
                    m_l.append(str(row.value))
                    row= cursor.fetchone()
                m_l.reverse()
                m_l1.extend(m_l)
                if(len(m_l1)<7):
                    n = 7-len(m_l1)
                    for i in range(0,n):
                        m_l1.append("None")
                print(m_l1)
                temp_metric = temp_metric.append(pd.Series(m_l1, index=['Name',str(par_year[0])+'Q1',str(par_year[1])+'Q1',str(par_year[2])+'Q1',str(par_year[3])+'Q1',str(par_year[4])+'Q1',str(par_year[5])+'Q1']), ignore_index=True) 
                types = ["Analyst metric","Database metric"]
                temp_metric['Type'] = types
                #temp_metric=temp_metric.set_index('Type')
                #display(temp_metric)
                print(tabulate(temp_metric, headers='keys', tablefmt='psql'))
                #print(m_l)
                #print(f"{'DB metric name and subsidiary id =':<30}{poss[0]:>15}",
                      #f"\n{'Analyst metric name =':<30}{poss[1]:>15}",
                      #f"\n{'Row Number in excel =':<30}{poss[2]:>15}",
                      #f"\n{'Row Number in excel =':<30}{poss[3]:>15}")
                print("DB metric name and subsidiary id: \n" + poss[0])
                print("Analyst metric name: \n" + poss[1])
                print("Row Number in excel: \n" + poss[2])
                print("Subsidiary Name from Analyst: \n" + poss[3])
                #matched_df = matched_df.append(excel_data.loc[[int(poss[2])],col[0:7]], ignore_index = True)
            print()
            #print("The definately matched metrics are")
            print()
            #display(matched_df)
            #print("The matches made are")
            #for m in def_metrics:
            #   print(m)
            #display(not_present)
    
        
        cnt =0
        l = []
        for i in def_metrics:
            p = []
            p = i.split("--->")
            l.append(p)
        final_db = pd.DataFrame(l,columns =['Database_metric','Analyst_metric','Row Number','Subsidiary Name'])
        d = list(final_db['Analyst_metric'])
        A= []
        for i in d:
            if d.count(i)>1 and i not in A:
                #display(final_db.loc[final_db['Analyst_metric']==i])
                print(tabulate(final_db.loc[final_db['Analyst_metric']==i], headers='keys', tablefmt='psql'))
                inp=input("\nEnter the rows you want to make the equation with, seperated by a comma (,). If you do not want to make an equation, press N")
                if inp is 'N':
                    l = list(final_db.loc[final_db['Analyst_metric']==i].index)
                    final_db = final_db.drop(l, axis =0)
                    final_db.reset_index(drop = True)
                else:
                    inp1 = inp.split(",")
                    for j in range(0,len(inp1)):
                        inp1[j] = int(inp1[j])
                    l = list(final_db.loc[final_db['Analyst_metric']==i].index)
                    for k in inp1:
                        l.remove(k)
                    final_db = final_db.drop(l,axis=0)
                    final_db.reset_index(drop = True)
                A.append(i)
        #display(final_db)
        final_flag ='N'
        print("\n\nThe list of matches to be thrown into the database are: ")
        final_db = final_db[['Analyst_metric','Database_metric','Row Number','Subsidiary Name']]
        while final_flag is not 'Y':
            #display(final_db)
            print(tabulate(final_db, headers='keys', tablefmt='psql'))
            inp = input("If you are happy with these values to be thrown into the database, type [Y]. If not, type the row numbers seperated by a comma (,)\n")
            if inp is 'Y':
                final_flag = 'Y'
                continue
            else:
                fin =[]
                fin = inp.split(",")
            for i in range(0,len(fin)):
                fin[i] = int(fin[i])
            final_db = final_db.drop(fin,axis = 0)    
    if got_cor:
        return 0
    else:
        return 1
   
    '''qw=input()
            if qw=='q':
                sys.exit()'''


# In[32]:



#get the minimum year present in dtabase for given ticker
def min_year(ticker):
    select_stmt1 = F" select distinct financial_year from dbo.cc_actual_metrics_consolidated where ticker_code={ticker} order by financial_year"
    cursor.execute(select_stmt1)
    row = cursor.fetchone()
    res_array = []
    while row:
        res_array.append(row[0])
        row = cursor.fetchone()
    #print("Minimum year present in the data base for the given ticker: ")    
    #print(res_array[0])
    return(res_array[0])


# In[ ]:





# In[33]:


#@safe_run
def get_metric_name (year,ticker,value,a_metric,quarter='Q1'):
    flag = 0
    flag_not_found = 1
    possible_list = ""
    try:
        index = excel_data1[excel_data1['Metric/Sub Name']==a_metric].index.item()
        possible_list = str(excel_data1['Las Vegas Sands Corp.'][index])
        param = possible_list.strip("][").replace("'","").replace(", ",",").replace(" ,",",")
        flag = 1
    except:
        pass
    #Also write a conditional if statement for the case of an empty list 
    ticker_code = "'"+ticker+"'"
    year = year
    quarter = "'"+quarter+"'"
    if type(value)==int or (type(value)==float and str(value)!='nan'):
        if abs(value)>=1:
            if abs(value) in [1,2,3,4,5,6,7,8,9] and type(value) is not float:
                valueo= "'"+str(value)+"'" #orginal value
                value3 = "'"+str(-1*int(value))+".%"+"'" #integered with negation
                if(len(possible_list)>2 and flag==1):
                    select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                         F"metric_name in (select id from [adm_fnSplitter]('{param}')) and "                         F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                         F"(value like {valueo} or value like {value3})"   
                else:
                    select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                         F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                         F"(value like {valueo} or value like {value3})"
                subs_list = []    
                metric_list=[]
                cursor.execute(select_stmt1)
                row = cursor.fetchone()
                while row:
                    flag_not_found = 0
                    metric_list.append(row.metric_name)
                    subs_list.append(row.subsidiary_id)
                    row = cursor.fetchone()
            else:
            #when value is greater than 1
                valueo= "'"+str(value)+"'" #orginal value
                value1 = "'"+str(int(value))+".%"+"'" # integered value
                value2 = "'"+str(round(value, 2))+"%"+"'" #value when rounded upto 2
                value3 = "'"+str(-1*int(value))+".%"+"'" #integered with negation
                value4 = "'"+str(-1*round(value,2))+"%"+"'" # round upto 2 with negation
                #print(valueo,value1,value2,value3,value4)
                if(len(possible_list)>2 and flag==1):
                    select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                         F"metric_name in (select id from [adm_fnSplitter]('{param}')) and "                         F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                         F"(value like {valueo} or value like {value1} or value like {value2} or value like {value3} or value like {value4})"   
                else:
                    select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                         F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                         F"(value like {valueo} or value like {value1} or value like {value2} or value like {value3} or value like {value4})"
                subs_list = []    
                metric_list=[]
                cursor.execute(select_stmt1)
                row = cursor.fetchone()
                while row:
                    flag_not_found = 0
                    metric_list.append(row.metric_name)
                    subs_list.append(row.subsidiary_id)
                    row = cursor.fetchone()
        else:
            #if value less than 1
            valueo= "'"+str(value*100)+"'" #orginal with 100 multiplied
            value1 = "'"+str(int(value*100))+".%"+"'" # multiply by 100 and then integered
            value3 = "'"+str(-1*int(value*100))+".%"+"'" # multiply by 100 and integer and negate it
            #print("with 100",valueo,value1,value3)
            if(len(possible_list)>2 and flag==1):
                select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                     F"metric_name in (select id from [adm_fnSplitter]('{param}')) and "                     F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                     F"(value like {valueo} or value like {value1} or value like {value3})"
            else:
                select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                     F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                     F"(value like {valueo} or value like {value1} or value like {value3})"
            subs_list = []    
            metric_list=[]
            cursor.execute(select_stmt1)
            row = cursor.fetchone()
            flag_got_with_100=0
            while row:
                metric_list.append(row.metric_name)
                subs_list.append(row.subsidiary_id)
                row = cursor.fetchone()
                flag_got_with_100=1

            #if not flag_got_with_100:
            valueo = "'"+str(value)+"'" #orginal value
            value1 = "'"+str(round(value,2))+"%"+"'" # round upto 2
            value3 = "'"+str(-1*round(value,2))+"%"+"'" # round upto 2  and negate it
            #print("without 100",valueo,value1,value3)
            if(len(possible_list)>2 and flag==1):
                select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                     F"metric_name in (select id from [adm_fnSplitter]('{param}')) and "                     F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                     F"(value like {valueo} or value like {value1} or value like {value3})"
            else:
                select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                     F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                     F"(value like {valueo} or value like {value1} or value like {value3})"
            cursor.execute(select_stmt1)
            row = cursor.fetchone()
            while row:
                #to_confirm
                metric_list.append(row.metric_name)
                subs_list.append(row.subsidiary_id)
                row = cursor.fetchone()
                flag_not_found = 0
        if(flag_not_found == 1 and flag == 1):
            if abs(value)>=1:
                if abs(value) in [1,2,3,4,5,6,7,8,9] and type(value) is not float:
                    valueo= "'"+str(value)+"'" #orginal value
                    value3 = "'"+str(-1*int(value))+".%"+"'" #integered with negation
                    if(len(possible_list)>2 and flag==1):
                        select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                         F"metric_name in (select id from [adm_fnSplitter]('{param}')) and "                         F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                         F"(value like {valueo} or value like {value3})"   
                    else:
                        select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                         F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                         F"(value like {valueo} or value like {value3})"
                    subs_list = []    
                    metric_list=[]
                    cursor.execute(select_stmt1)
                    row = cursor.fetchone()
                    while row:
                        flag_not_found = 0
                        metric_list.append(row.metric_name)
                        subs_list.append(row.subsidiary_id)
                        row = cursor.fetchone()
                else:        
                #when value is greater than 1
                    valueo= "'"+str(value)+"'" #orginal value
                    value1 = "'"+str(int(value))+".%"+"'" # integered value
                    value2 = "'"+str(round(value, 2))+"%"+"'" #value when rounded upto 2
                    value3 = "'"+str(-1*int(value))+".%"+"'" #integered with negation
                    value4 = "'"+str(-1*round(value,2))+"%"+"'" # round upto 2 with negation
                    select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                     F"metric_name not in (select id from [adm_fnSplitter]('{param}')) and "                     F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                     F"(value like {valueo} or value like {value1} or value like {value2} or value like {value3} or value like {value4})"
                    metric_list=[]
                    subs_list = []
                    cursor.execute(select_stmt1)
                    row = cursor.fetchone()
                    while row:
                        metric_list.append(row.metric_name)
                        subs_list.append(row.subsidiary_id)
                        row = cursor.fetchone()
            else:
                #if value less than 1
                valueo= "'"+str(value*100)+"'" #orginal with 100 multiplied
                value1 = "'"+str(int(value*100))+".%"+"'" # multiply by 100 and then integered
                value3 = "'"+str(-1*int(value*100))+".%"+"'" # multiply by 100 and integer and negate it  
                select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                     F"metric_name not in (select id from [adm_fnSplitter]('{param}')) and "                     F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                     F"(value like {valueo} or value like {value1} or value like {value3})"
                #print("with 100",valueo,value1,value3)
                metric_list=[]
                subs_list = []
                cursor.execute(select_stmt1)
                row = cursor.fetchone()
                flag_got_with_100=0
                while row:
                    metric_list.append(row.metric_name)
                    subs_list.append(row.subsidiary_id)
                    row = cursor.fetchone()
                    flag_got_with_100=1
                valueo = "'"+str(value)+"'" #orginal value
                value1 = "'"+str(round(value,2))+"%"+"'" # round upto 2
                value3 = "'"+str(-1*round(value,2))+"%"+"'" # round upto 2  and negate it
                select_stmt1 = F"select distinct metric_name, subsidiary_id from dbo.cc_actual_metrics_consolidated where "                     F"metric_name not in (select id from [adm_fnSplitter]('{param}')) and "                     F"ticker_code={ticker_code} and financial_year = {year} and quarter = {quarter} and "                     F"(value like {valueo} or value like {value1} or value like {value3})"   
                #print("without 100",valueo,value1,value3)
                cursor.execute(select_stmt1)
                row = cursor.fetchone()
                while row:
                    #to_confirm
                    metric_list.append(row.metric_name)
                    subs_list.append(row.subsidiary_id)
                    row = cursor.fetchone()
        return(metric_list,subs_list)
    else:
        return([],[])


# In[34]:


#function which tells about the quater and year the particular column belongs to
def about_q_year(checkar):
    
    checkar_duplicate=np.copy(checkar)
    quater=''
    year=''
    year_flag=0
    quater_flag=0
    q4_check=0
    checkar=[]
    datetime_flag=1
    for i_che in checkar_duplicate:
        checkar.append(str(i_che))
    
    for i_che in range(0,len(checkar)):
        
        if not year_flag:
            #checking if year is presenr in (2017,2016 format)
            for j_y in year_pattern_int:
                if str(j_y) in checkar[i_che]:
                    year=str(j_y)[2:]
                    checkar[i_che]=checkar[i_che].replace(str(j_y),'')
                    checkar[i_che]=checkar[i_che].replace(year,'')
                    year_flag=1
                    break
        if not year_flag:
            #checking if the year is present in 2 digited format like 17,16)
            for j_y in year_pattern_2di:
                if str(j_y) in checkar[i_che]:
                    year=str(j_y)
                    checkar[i_che]=checkar[i_che].replace(year,'')
                    year_flag=1
                    break

        if not quater_flag:
            # getting info about the quater
            for i_p in quater_pattern:
                checkar[i_che]=checkar[i_che].replace(year,'')
                checkar[i_che]=checkar[i_che].replace("20"+year,'')
                if i_p in checkar[i_che].lower():
                    quater=i_p
                    quater_flag=1
                    break       
        
        if year_flag and quater_flag:
            break
    #returning the year and quater and if its 4th quatered or not
    return(year,quater,q4_check)


# In[35]:


# funtion to avoid te rough work present in top and retutn the row no from which its important
def startingrow(wb,sheet_name):
    curr_sheet = wb[sheet_name]
    flag = 0
    row = 0
    for col in curr_sheet.iter_cols(max_col=20,max_row=15):
        for cell in col:
            if cell.value is not None:
                for data in check_list:
                    if fuzz.partial_ratio(str(data),str(cell.value)) == 100:
                        row = cell.row
                        flag =1
                        break
                if flag == 1:
                    break
        if flag == 1:
            break

    #print(row)
    return(row)


# In[36]:


#function which return all the metric anme with repect to ticker in actuals
#@safe_run
def get_metriclist_from_db(year, ticker):

    ticker_code = "'"+ticker+"'"
    year = year
    select_stmt1 = F"select distinct metric_name from dbo.cc_actual_metrics_consolidated where ticker_code = {ticker_code} "         F" and financial_year = {year}"

    cursor.execute(select_stmt1)
    row = cursor.fetchone()
    metric_list_comp = []
    while row:
        metric_list_comp.append(row[0])
        row = cursor.fetchone()

    return (metric_list_comp)


# In[37]:


#function which return aall the metric name wrt to analyst from estimated database
@safe_run
def get_metriclist_wrt_analyst(ticker,analyst):
    ticker_code = "'"+ticker+"'"
    analyst_name="'%"+analyst+"%'"
    select_stmt1 = F" select distinct metric_name from [dbo].[cc_metrics_consolidated] where ticker_code = {ticker_code} and quarter='q1' and analyst_name like {analyst_name} order by metric_name"
    cursor.execute(select_stmt1)
    row = cursor.fetchone()
    res_array = []
    while row:
        res_array.append(row[0])
        row = cursor.fetchone()
    return(res_array)


# In[38]:


ind = 0
for i in excel_files:
    print(str(ind) + ") "+ i[28:])
    ind = ind+1
inpu = input("\n Enter the serial number of the file you would like to process\n")    
filen=excel_files[int(inpu)]
print(filen)
path_list=formating(filen)    
print("\n The current ticker that is detected is ", path_list[3])
mp = opx.load_workbook("E:\\final\\Corpus\\LVS Mapping ALL.xlsx",data_only=True)
sheets = mp.sheetnames
if(path_list[4] in sheets):
    chek=path_list[4]
else:
    count = 0
    for sheet in sheets:
        print(str(count) + ") " + str(sheet))
        count = count + 1
    ipn = input("\nenter the number of the sheet you want to use ")
    ipn = int(ipn)
    chek = sheets[ipn]
excel_data1 = pd.read_excel("E:\\final\\Corpus\\LVS Mapping ALL.xlsx",sheet_name=chek) 
bscf_names = list(set(list(excel_data1['Worksheet'])))
a=list(excel_data1['Metric/Sub Name'])
A= []
for i in range(len(a)):
    if a.count(a[i])>1 and a[i] not in A:
        dup = excel_data1.loc[excel_data1['Metric/Sub Name']==a[i]]
        ms = list(dup['Las Vegas Sands Corp.'])
        excel_data1.at[i,'Las Vegas Sands Corp.'] = ms
    A.append(a[i])   
excel_data1 = excel_data1.drop_duplicates(subset='Metric/Sub Name', keep = 'first').reset_index(drop=True)
#display(excel_data1)
val=function1(filen,path_list)
if val==1:
    print("FAULT")
elif val==2:
    print("fault with sheet name")


# In[ ]:




